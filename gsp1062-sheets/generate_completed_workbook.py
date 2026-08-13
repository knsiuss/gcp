#!/usr/bin/env python3
"""
Generate complete Google Sheets XLSX file for GSP1062
"""

import urllib.request
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule, FormulaRule

def main():
    url = "https://storage.googleapis.com/cloud-training/GoogleSheets/On%20The%20Rise%20Bakery%20Customers%20and%20Items.xlsx"
    file_path = "On_The_Rise_Bakery_Completed.xlsx"
    urllib.request.urlretrieve(url, "temp_orig.xlsx")

    wb = openpyxl.load_workbook("temp_orig.xlsx")

    # =========================================================================
    # 1. PROCESS 'Items' SHEET
    # =========================================================================
    ws_items = wb["Items"]

    # Freeze 1st row
    ws_items.freeze_panes = "A2"

    # Insert column B for Formatted Name
    ws_items.insert_cols(2)
    ws_items.cell(row=1, column=2, value="Formatted Name")

    # Update headers
    ws_items.cell(row=1, column=1, value="Item Name")
    ws_items.cell(row=1, column=3, value="Unit Price")
    ws_items.cell(row=1, column=4, value="Number of Items")
    ws_items.cell(row=1, column=5, value="Total Cost")
    ws_items.cell(row=1, column=6, value="Truncated Unit Prices")

    # Apply formulas for Items (rows 2 to 15)
    max_row_items = ws_items.max_row
    for r in range(2, max_row_items + 1):
        item_val = ws_items.cell(row=r, column=1).value
        if item_val:
            ws_items.cell(row=r, column=2, value=f"=PROPER(A{r})")
            ws_items.cell(row=r, column=5, value=f"=ROUND(C{r}*D{r}, 2)")
            ws_items.cell(row=r, column=5).number_format = '$#,##0.00'
            ws_items.cell(row=r, column=6, value=f"=TRUNC(C{r}, 4)")

    # Sort Items by Number of Items (Column 4 - D) ascending
    # We sort rows 2 to max_row_items based on Column 4 value
    item_rows = []
    for r in range(2, max_row_items + 1):
        vals = [ws_items.cell(row=r, column=c).value for c in range(1, 7)]
        if any(vals):
            item_rows.append(vals)

    item_rows.sort(key=lambda x: (x[3] if x[3] is not None else 0))

    for r_idx, row_vals in enumerate(item_rows, start=2):
        for c_idx, val in enumerate(row_vals, start=1):
            cell = ws_items.cell(row=r_idx, column=c_idx)
            cell.value = val
            if c_idx == 5:
                cell.number_format = '$#,##0.00'

    # =========================================================================
    # 2. CREATE SHEET 'Items Sorted By Unit Price'
    # =========================================================================
    if "Items Sorted By Unit Price" in wb.sheetnames:
        del wb["Items Sorted By Unit Price"]

    ws_sorted = wb.create_sheet(title="Items Sorted By Unit Price")
    ws_sorted.cell(row=1, column=1, value="=SORT(Items!A1:Items!C15, Items!B1:Items!B15, FALSE)")

    # =========================================================================
    # 3. PROCESS 'Customers' SHEET
    # =========================================================================
    ws_cust = wb["Customers"]

    # Header for Column D
    ws_cust.cell(row=1, column=4, value="Valid Email Address?")
    ws_cust.cell(row=1, column=7, value="Personalized Greeting")

    max_row_cust = ws_cust.max_row

    # Clean duplicates in Customers sheet based on Columns A to D
    seen_rows = set()
    cleaned_rows = []
    
    headers_cust = [ws_cust.cell(row=1, column=c).value for c in range(1, 5)]

    for r in range(2, max_row_cust + 1):
        first = ws_cust.cell(row=r, column=1).value
        last = ws_cust.cell(row=r, column=2).value
        email = ws_cust.cell(row=r, column=3).value
        
        if first or last or email:
            key = (str(first).strip() if first else "", str(last).strip() if last else "", str(email).strip() if email else "")
            if key not in seen_rows:
                seen_rows.add(key)
                cleaned_rows.append((first, last, email))

    # Clear existing rows in Customers
    ws_cust.delete_rows(2, max_row_cust)

    # Write cleaned rows
    for r_idx, (first, last, email) in enumerate(cleaned_rows, start=2):
        ws_cust.cell(row=r_idx, column=1, value=first)
        ws_cust.cell(row=r_idx, column=2, value=last)
        ws_cust.cell(row=r_idx, column=3, value=email)
        ws_cust.cell(row=r_idx, column=4, value=f"=ISEMAIL(C{r_idx})")
        ws_cust.cell(row=r_idx, column=7, value=f'=CONCATENATE("Hello ", TRIM(A{r_idx}),",")')

    # Add Conditional Formatting for duplicate emails on C1:C100
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws_cust.conditional_formatting.add("C1:C100", FormulaRule(formula=['COUNTIF(C:C,C1)>1'], fill=red_fill))

    wb.save(file_path)
    print(f"Successfully generated completed workbook: {file_path}")

if __name__ == "__main__":
    main()
